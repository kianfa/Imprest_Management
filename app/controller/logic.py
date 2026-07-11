from __future__ import annotations

import traceback

from openpyxl.utils import get_column_letter
from dataclasses import dataclass
import shutil
from app.data.data_base import Load_Save_Data,UserSession
from PyQt6.QtGui import QStandardItem, QImage, QAction, QIcon, QPainter, QPageLayout, QPageSize
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from typing import Optional, Union
from pathlib import Path
from PyQt6.QtWidgets import (
    QWidget, QMessageBox, QComboBox, QInputDialog, QMenu, QApplication, QTableView, QFileDialog, QRadioButton, QLineEdit
)
from app.ui.edit_record_dialog import EditRecordDialog
from PyQt6.QtCore import Qt, QSortFilterProxyModel, QSettings, QMarginsF, QEventLoop
from PyQt6.QtPrintSupport import QPrinter
import sys
import os


if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent.parent.parent

PathLike = Union[str, Path]



class receipt_entry_logic:
    def __init__(self):
        self.selected_image_paths = []

    def enhance_combo(self, combo: QComboBox, settings_key: str, default_items=None) -> None:
        if default_items is None:
            default_items = []

        combo.settings_key = settings_key
        combo._special_text = "➕ Add new…"
        combo._block_selection = False
        combo._default_items = default_items

        def load_items() -> None:
            settings = QSettings('MyCompany', 'MyApp')
            items = settings.value(combo.settings_key, [])
            if not items:
                items = combo._default_items.copy()

            combo.blockSignals(True)
            combo.clear()
            if items:
                combo.addItems(items)
            combo.addItem(combo._special_text)
            combo.blockSignals(False)

        def save_items() -> None:
            items = []
            for i in range(combo.count()):
                text = combo.itemText(i)
                if text != combo._special_text:
                    items.append(text)
            QSettings('MyCompany', 'MyApp').setValue(combo.settings_key, items)

        def on_index_changed(index) -> None:
            if combo._block_selection:
                return
            if combo.itemText(index) == combo._special_text:
                combo._block_selection = True
                new_item, ok = QInputDialog.getText(combo, "Add new item", "Enter item name:")
                if ok and new_item.strip():
                    new_item = new_item.strip()
                    if combo.findText(new_item) == -1:
                        insert_pos = combo.count() - 1  # before the special item
                        combo.insertItem(insert_pos, new_item)
                        save_items()
                        combo.setCurrentIndex(insert_pos)
                    else:
                        QMessageBox.information(combo, "Duplicate", f'"{new_item}" already exists.')
                # Reset selection to first normal item
                if combo.count() > 1:
                    combo.setCurrentIndex(0)
                combo._block_selection = False

        def remove_item_at_row(row) -> None:
            if row < 0 or row >= combo.count():
                return
            item_text = combo.itemText(row)
            if item_text == combo._special_text:
                return
            reply = QMessageBox.question(
                combo, "Remove item",
                f'Are you sure you want to remove "{item_text}"?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                combo._block_selection = True
                combo.removeItem(row)
                save_items()
                combo._block_selection = False

        def on_view_context_menu(point) -> None:
            """Called when right‑clicking inside the dropdown list."""
            view = combo.view()
            index = view.indexAt(point)
            if not index.isValid():
                return
            row = index.row()
            if combo.itemText(row) == combo._special_text:
                return
            menu = QMenu()
            remove_action = QAction("Remove", combo)
            icon = QIcon.fromTheme("edit-delete")
            if not icon.isNull():
                remove_action.setIcon(icon)
            else:
                remove_action.setText("🗑 Remove")
            remove_action.triggered.connect(lambda checked, r=row: remove_item_at_row(r))
            menu.addAction(remove_action)
            menu.exec(view.mapToGlobal(point))

        # Get the dropdown view (this creates it if necessary)
        view = combo.view()
        view.setContextMenuPolicy(Qt.ContextMenuPolicy.CustomContextMenu)
        view.customContextMenuRequested.connect(on_view_context_menu)

        # Disconnect any previous connection to avoid duplicates
        try:
            combo.currentIndexChanged.disconnect()
        except TypeError:
            pass
        combo.currentIndexChanged.connect(on_index_changed)

        # Attach helper methods (optional)
        combo.load_items = load_items
        combo.save_items = save_items

        load_items()

    def browse_image(
            self,
            parent: QWidget,
            title: str = "add image",
            start_dir: Optional[str | Path] = None,
        ) -> list[str]:
        start_dir_str = str(start_dir) if start_dir else ""

        file_paths, _ = QFileDialog.getOpenFileNames(
            parent,
            title,
            start_dir_str,
            "Images (*.png *.jpg *.jpeg *.bmp *.gif *.webp);;All Files (*)",
        )

        return file_paths


    def add_image_logic(self, parent) -> None:
        new_paths = self.browse_image(parent=parent, title="Add images", start_dir="./")
        if not new_paths:
            return

        # initialize if not present
        if not hasattr(self, "selected_image_paths"):
            self.selected_image_paths = []

        # add without duplicates
        for p in new_paths:
            if p not in self.selected_image_paths:
                self.selected_image_paths.append(p)

    def show_field_error(self, field_name) -> None:
        QMessageBox.warning(None,"Input Error", f"Field '{field_name}' cannot be empty.")

    def show_duplicate_error(self, invoice_number) -> None:
        QMessageBox.warning(None, "Duplicate Invoice",
                            f"'{invoice_number}' already exists.\nPlease use a unique number.")

    def show_wrong_type_error(self, field_name) -> None:
        QMessageBox.warning(None, "Wrong Type", f"Field '{field_name}' is not a number.")

    def duplicate_check_invoice(self, number) -> bool:
        db = Load_Save_Data()
        if db.invoice_exists(number):
            return False
        return True

    def duplicate_check_project(self, number) -> bool:
        db = Load_Save_Data()
        if db.project_exists(number):
            return False
        return True

from app.data.data_base import DataBase

@dataclass
class LoginResult:
    ok: bool
    role: str = ""
    error_message: str = ""

class main_window_logic:
    @staticmethod
    def validate_inputs(username: str, password: str) -> LoginResult:
        if not username:
            return LoginResult(False, error_message="Username cannot be empty.")
        if not password:
            return LoginResult(False, error_message="Password cannot be empty.")
        return LoginResult(True)

    @classmethod
    def login(cls, username: str, password: str) -> LoginResult:
        # Step 1: validate input
        validation = cls.validate_inputs(username, password)
        if not validation.ok:
            return validation

        # Step 2: verify against database
        success, role = DataBase.verify_login(username, password)
        if success:
            return LoginResult(True, role)
        else:
            return LoginResult(False, error_message="Incorrect username or password.")




class calling_page_logic:
    def __init__(self) ->None:
        super().__init__()

    def load_invoices(self,
                      rbi : QRadioButton, lei : QLineEdit,
                      rbp : QRadioButton, lep: QLineEdit,
                      rbt : QRadioButton, lesd : QLineEdit, leed : QLineEdit,
                      rbrd : QRadioButton, lerd : QLineEdit,
                      lee : QLineEdit) -> None:
        # Your repo functions return (headers, rows), so we ignore headers
        if rbi.isChecked():
            rows = Load_Save_Data.get_invoices_by_Invoice_NO(lei.text().strip())

        elif rbp.isChecked():
            rows = Load_Save_Data.get_invoices_by_Project_Code(lep.text().strip())

        elif rbt.isChecked():
            rows = Load_Save_Data.get_invoices_by_time_range(lesd.text().strip(), leed.text().strip())

        elif rbrd.isChecked():
            rows = Load_Save_Data.get_invoices_by_regestrationdate(lerd.text().strip())

        else:  # rbExplanation
            rows = Load_Save_Data.get_invoices_by_explanation(lee.text().strip())

        self.model.clear()
        self.model.setColumnCount(len(self.headers))
        self.model.setHorizontalHeaderLabels(self.headers)  # real header, real labels

        # Data rows
        for row in rows:
            items = []
            for v in row:
                if isinstance(v, float):
                    item = QStandardItem(f"{int(v):,}")
                else:
                    item = QStandardItem("" if v is None else str(v))
                item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                items.append(item)
            self.model.appendRow(items)

        self.tableView.setColumnHidden(0, True)
        # Default sort by record_date
        date_col_index = self.headers.index("record_date")
        self.tableView.sortByColumn(date_col_index, Qt.SortOrder.AscendingOrder)

    def edit_record(self, table: QTableView) -> None:
        rows = table.selectionModel().selectedRows()
        if len(rows) != 1:
            QMessageBox.warning(
                None,
                "Selection Error",
                "Please select exactly one record to edit."
            )
            return

        row = rows[0].row()
        record_id = self.model.item(row, 0).text()
        created_by = self.model.item(row, 9).text()

        if UserSession.role != 'admin' and created_by != UserSession.full_name:
            QMessageBox.warning(None, "Permission Denied", "You can only edit your own records.")
            return

        record_data = Load_Save_Data().get_record_by_id(record_id)
        if not record_data:
            QMessageBox.critical(None, "Error", "Record not found.")
            return

        dialog = EditRecordDialog(record_data)
        if dialog.exec():
            updated = dialog.get_updated_data()

            QMessageBox.information(None, "Edited", "Record has been edited.")
            Load_Save_Data.update_record(record_id, updated)

    def delete_record(self, table: QTableView) -> None:
        # Get selected row
        selected = table.selectedIndexes()
        if not selected:
            QMessageBox.warning(None, "No Selection", "Please select a record to delete.")
            return

        row = selected[0].row()
        record_id = self.model.item(row, 0).text()
        created_by = self.model.item(row, 8).text()

        # Permission check
        if UserSession.role != 'admin' and created_by != UserSession.full_name:
            QMessageBox.warning(None, "Permission Denied", "You can only delete your own records.")
            return

        # Confirmation dialog
        reply = QMessageBox.question(None, "Confirm Delete",
                                     "This action cannot be undone. Do you want to continue?",
                                     QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No)

        target_path = Path(__file__).parent.parent.parent / "image_records" / str(record_id)

        if reply == QMessageBox.StandardButton.Yes:
            if target_path.exists():
                shutil.rmtree(target_path)

            Load_Save_Data.soft_delete_record(record_id)
            QMessageBox.information(None, "Deleted", "Record has been deleted.")

    def filtering(self, table: QTableView, text: str) -> None:
        if not hasattr(self, 'proxy'):
            self.proxy = QSortFilterProxyModel()
            self.proxy.setSourceModel(table.model())
            self.proxy.setFilterCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            self.proxy.setFilterKeyColumn(8)
            self.proxy.filterAcceptsRow = lambda row, parent: True if row == 0 else QSortFilterProxyModel.filterAcceptsRow(self.proxy, row, parent)
            table.setModel(self.proxy)

        self.proxy.setFilterFixedString(text)


class exporting:
    def __init__(self) -> None:
        super().__init__()

    def export_table_to_pdf(self, table, file_path):
        if getattr(sys, 'frozen', False):
            base = Path(sys.executable).parent
        else:
            base = Path(__file__).resolve().parent.parent.parent

        export_view = QTableView()
        export_view.setModel(table.model())
        export_view.setFont(table.font())
        export_view.setStyleSheet(table.styleSheet())

        for col in range(table.model().columnCount()):
            export_view.setColumnWidth(col, table.columnWidth(col))

        printer = QPrinter(QPrinter.PrinterMode.ScreenResolution)
        printer.setOutputFormat(QPrinter.OutputFormat.PdfFormat)
        printer.setOutputFileName(file_path)
        printer.setPageSize(QPageSize(QPageSize.PageSizeId.A4))
        printer.setPageOrientation(QPageLayout.Orientation.Landscape)
        printer.setPageMargins(QMarginsF(10, 10, 10, 10))

        painter = QPainter()
        if not painter.begin(printer):
            raise RuntimeError("Could not begin painter")

        page_rect = printer.pageRect(QPrinter.Unit.DevicePixel)
        page_count = 1
        painter.drawText(
            int(page_rect.width() - 5),  # x: from left
            int(page_rect.height() - 1),  # y: from top
            str(page_count)
        )

        try:
            export_view.resizeColumnsToContents()
            export_view.resizeRowsToContents()
            export_view.setFixedHeight(
                export_view.horizontalHeader().height()
                + export_view.verticalHeader().length() + 4
            )
            export_view.setFixedWidth(
                export_view.verticalHeader().width()
                + export_view.horizontalHeader().length() + 4
            )
            QApplication.processEvents(QEventLoop.ProcessEventsFlag.ExcludeUserInputEvents)
            export_view.setColumnHidden(0, True)

            scale = 1.35
            painter.save()
            painter.scale(scale, scale)
            export_view.render(painter)
            painter.restore()

            row_count = table.model().rowCount()
            model = table.model()

            for row in range(1, row_count):
                printer.newPage()
                page_count+=1
                painter.drawText(
                    int(page_rect.width() - 5),  # x: from left
                    int(page_rect.height() - 1),  # y: from top
                    str(page_count)
                )

                export_view.setUpdatesEnabled(False)
                for r in range(row_count):
                    export_view.setRowHidden(r, True)
                export_view.setRowHidden(0, False)
                export_view.setColumnHidden(0, True)
                export_view.setRowHidden(row, False)
                export_view.setUpdatesEnabled(True)

                export_view.setFixedHeight(
                    export_view.horizontalHeader().height()
                    + export_view.rowHeight(0)
                    + export_view.rowHeight(row) + 10
                )
                QApplication.processEvents(QEventLoop.ProcessEventsFlag.ExcludeUserInputEvents)

                id = str(model.data(model.index(row, 0))).strip()
                images_path = base / "image_records" / id

                image_files = []
                if images_path.exists() and images_path.is_dir():
                    image_files = [
                        f for f in os.listdir(images_path)
                        if f.lower().endswith((".png", ".jpg", ".jpeg", ".bmp", ".webp"))
                    ]

                painter.save()
                painter.scale(scale, scale)
                export_view.render(painter)
                painter.restore()

                table_height = export_view.height() * scale
                x, y = 50, int(table_height)
                page_height = printer.pageRect(QPrinter.Unit.DevicePixel).height()

                for file in image_files:
                    image = QImage(str(images_path / file))
                    scaled = image.scaled(
                        700, 700,
                        Qt.AspectRatioMode.KeepAspectRatio,
                        Qt.TransformationMode.SmoothTransformation
                    )
                    if y + scaled.height() > page_height:
                        printer.newPage()
                        page_count+=1
                        painter.drawText(
                            int(page_rect.width() - 5),  # x: from left
                            int(page_rect.height() - 10), # y: from top
                            str(page_count)
                        )
                        painter.save()
                        painter.scale(scale, scale)
                        export_view.render(painter)  # export_view still has this row visible
                        painter.restore()

                        table_height = export_view.height() * scale
                        y = int(table_height) + 5

                    painter.drawImage(x, y, scaled)
                    y += scaled.height() + 4

        finally:
            painter.end()
            export_view.deleteLater()


    def export_tableview_to_excel(self, view) -> None:
        model = view.model()
        header = view.horizontalHeader()

        if model is None:
            raise RuntimeError("Could not find model")

        # Get visible columns in display order
        export_col = []
        for visual in range(header.count()):
            logical = header.logicalIndex(visual)
            if 0 <= logical < model.columnCount():
                export_col.append(logical)

        if not export_col:
            raise RuntimeError("Could not find columns")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Export Table"

        header_font = Font(bold=True)
        header_align = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

        # ==================================================
        # HEADER ROW (ROW 1)
        # ==================================================

        # Numbering column header
        cell = ws.cell(row=1, column=1, value="No.")
        cell.font = header_font
        cell.alignment = header_align

        # Table headers
        for excel_col, logical_col in enumerate(export_col, start=2):
            title = model.headerData(
                logical_col,
                Qt.Orientation.Horizontal,
                Qt.ItemDataRole.DisplayRole
            )

            text = "" if title is None else str(title)

            cell = ws.cell(
                row=1,
                column=excel_col,
                value=text
            )
            cell.font = header_font
            cell.alignment = header_align

        ws.freeze_panes = "A2"

        # ==================================================
        # DATA ROWS (START FROM ROW 2)
        # ==================================================

        rows = model.rowCount()

        for r in range(rows):
            excel_row = r + 1

            # Row number
            ws.cell(
                row=excel_row + 1,
                column=1,
                value=r + 1
            )

            # Data columns
            for excel_col, logical_col in enumerate(export_col, start=2):
                idx = model.index(r, logical_col)

                value = model.data(
                    idx,
                    Qt.ItemDataRole.DisplayRole
                )

                if value is None:
                    state = model.data(
                        idx,
                        Qt.CheckStateRole
                    )

                    if state is not None:
                        value = "✓" if state == Qt.Checked else "✗"
                    else:
                        value = ""

                ws.cell(
                    row=excel_row,
                    column=excel_col,
                    value=str(value)
                )

        # ==================================================
        # AUTO WIDTH
        # ==================================================

        total_columns = len(export_col) + 1

        for col in range(1, total_columns + 1):
            max_length = 0

            for row in range(1, ws.max_row + 1):
                value = ws.cell(row=row, column=col).value

                if value is not None:
                    max_length = max(
                        max_length,
                        len(str(value))
                    )

            letter = get_column_letter(col)

            if col == 1:
                ws.column_dimensions[letter].width = 8
            else:
                ws.column_dimensions[letter].width = max(
                    15,
                    min(max_length + 2, 60)
                )

        # ==================================================
        # SAVE FILE
        # ==================================================

        parent = view.window()

        path, _ = QFileDialog.getSaveFileName(
            parent,
            "Save Excel",
            "Export.xlsx",
            "Excel Workbook (*.xlsx)"
        )

        if not path:
            return

        if not path.lower().endswith(".xlsx"):
            path += ".xlsx"

        try:
            wb.save(path)

        except PermissionError:
            QMessageBox.warning(
                self,
                "Save failed",
                "Could not write the file.\n"
                "If it's open in Excel, close it and try again."
            )

        except OSError as e:
            QMessageBox.critical(
                self,
                "Save failed",
                f"OS error while saving:\n{e}"
            )

        except Exception as e:
            QMessageBox.critical(
                self,
                "Save failed",
                f"Unexpected error:\n{e}"
            )